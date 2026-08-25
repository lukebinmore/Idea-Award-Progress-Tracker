from openpyxl import load_workbook
from IAPT.core.config import load_config
from IAPT.core.database import upsert_students, upsert_results, upsert_schedule
from IAPT.core.exceptions import IAPTError
import logging

logger = logging.getLogger(__name__)


def read_excel_data(file_path, headers):
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()

    except FileNotFoundError as error:
        raise IAPTError(message="The selected file could not be found", error=error, file_path=file_path)
    except PermissionError:
        raise IAPTError(
            message="You do not have permission to open the selected file", error=error, file_path=file_path
        )
    except Exception as error:
        logger.exception("Unknown error")
        raise IAPTError(message="An Unknown Error has Occurred", error=error, file_path=file_path)

    if not rows:
        raise IAPTError(message="No data was found in the selected spreadsheet", extra={"file_path", file_path})

    spreadsheet_headers = rows[0]
    indexes = []

    for header in headers:
        if header not in spreadsheet_headers:
            raise IAPTError(
                message="One or more required columns missing from the selected spreadsheet", file_path=file_path
            )

        indexes.append(spreadsheet_headers.index(header))

    data = []
    for row in rows[1:]:
        data.append({header: row[index] for header, index in zip(headers, indexes)})

    return data


def import_students(file_path, class_name):
    try:
        logger.info("Starting students import")

        config = load_config("student_import")
        data = read_excel_data(file_path, config.values())

        students = []
        for row_number, student in enumerate(data, start=2):
            if not student[config["student_id"]]:
                logger.warning(
                    "Incomplete student data", extra={"row": row_number, "missing_data": config["student_id"]}
                )
                continue
            if not student[config["first_name"]]:
                logger.warning(
                    "Incomplete student data", extra={"row": row_number, "missing_data": config["first_name"]}
                )
                continue
            if not student[config["last_name"]]:
                logger.warning(
                    "Incomplete student data", extra={"row": row_number, "missing_data": config["last_name"]}
                )
                continue

            students.append(
                {
                    "student_id": student[config["student_id"]],
                    "first_name": student[config["first_name"]],
                    "last_name": student[config["last_name"]],
                }
            )

        upsert_students(students, class_name)

        logger.success(f"{len(students)} students processed")
    except IAPTError as error:
        logger.error("Students import failed", extra={"error": error})


def import_results(file_path):
    try:
        logger.info("Starting results import")

        config = load_config("results_import")
        data = read_excel_data(file_path, config.values())

        results = []
        for row in data:
            if row[config["badge_list"]]:
                row[config["badge_list"]] = row[config["badge_list"]].split(",")

            results.append(
                {
                    "student_id": row[config["student_id"]].split("@")[0],
                    "bronze_points_total": row[config["bronze_points_total"]] or 0,
                    "bronze_citizen": row[config["bronze_citizen"]] or 0,
                    "bronze_worker": row[config["bronze_worker"]] or 0,
                    "bronze_maker": row[config["bronze_maker"]] or 0,
                    "bronze_entrepreneur": row[config["bronze_entrepreneur"]] or 0,
                    "silver_points_total": row[config["silver_points_total"]] or 0,
                    "badge_list": row[config["badge_list"]] or [],
                }
            )

        upsert_results(results)

        logger.success(f"{len(results)} results processed")
    except IAPTError as error:
        logger.error("Results import failed", extra={"error": error})


def import_schedule(file_path):
    try:
        logger.info("Starting schedule import")

        config = load_config("schedule_import")
        data = read_excel_data(file_path, config.values())

        homeworks = []
        for homework in data:
            homework[config["due_date"]] = homework[config["due_date"]].date()

            homeworks.append(
                {
                    "badge_name": homework[config["badge_name"]] or "",
                    "category": homework[config["category"]] or "",
                    "points": homework[config["points"]],
                    "due_date": homework[config["due_date"]],
                }
            )

        upsert_schedule(homeworks)

        logger.success(f"{len(homeworks)} homeworks processed")
    except IAPTError as error:
        logger.error("Schedule import failed", extra={"error": error})
